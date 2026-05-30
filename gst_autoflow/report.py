"""
Excel Report Generator — color-coded, client-ready output.
Returns (bytes, filename) — never writes to disk.
"""
from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Colour palette
GREEN  = "C6EFCE"   # matched clean
YELLOW = "FFEB9C"   # amount mismatch
RED    = "FFC7CE"   # missing in GSTR-2A (ITC at risk)
BLUE   = "BDD7EE"   # extra in GSTR-2A, not in books
HEADER = "1F3864"   # dark navy header


def _header_style(ws, row: int = 1) -> None:
    for cell in ws[row]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _fill_rows(ws, fill_color: str, start_row: int = 2) -> None:
    fill = PatternFill("solid", fgColor=fill_color)
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
        for cell in row:
            cell.fill = fill


def _autofit(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def _write_sheet(wb, title: str, df: pd.DataFrame, fill_color: str, note: str | None = None) -> None:
    ws = wb.create_sheet(title=title)
    if note:
        ws.append([f"\u2139  {note}"])
        ws["A1"].font = Font(italic=True, color="595959")
        ws.append([])
        start_data = 3
    else:
        start_data = 1

    if df.empty:
        ws.append(["No records in this category."])
        return

    headers = list(df.columns)
    ws.append(headers)
    _header_style(ws, row=start_data)

    for _, row in df.iterrows():
        ws.append([row[c] for c in headers])

    _fill_rows(ws, fill_color, start_row=start_data + 1)
    _autofit(ws)


def _save(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_report(results: dict) -> tuple[bytes, str]:
    """GSTR-2A reconciliation report. Returns (bytes, filename)."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"GST_Reconciliation_Report_{ts}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    s = results["summary"]

    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum.sheet_view.showGridLines = False

    summary_data = [
        ["GST RECONCILIATION REPORT", ""],
        [f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}", ""],
        ["", ""],
        ["Category", "Count / Value"],
        ["Purchase Register \u2014 Total Invoices", s["total_pr"]],
        ["GSTR-2A \u2014 Total Invoices", s["total_gstr2a"]],
        ["", ""],
        ["\u2705  Matched (Clean)", s["matched_clean"]],
        ["\u26a0\ufe0f  Amount Mismatch", s["amount_mismatch"]],
        ["\U0001f534  Missing in GSTR-2A (ITC at Risk)", s["missing_in_gstr2a"]],
        ["\U0001f535  In GSTR-2A but Not in Books", s["not_in_books"]],
        ["", ""],
        ["Purchase Register \u2014 Total Taxable (\u20b9)", f"\u20b9{s['pr_taxable_total']:,.2f}"],
        ["GSTR-2A \u2014 Total Taxable (\u20b9)", f"\u20b9{s['g2a_taxable_total']:,.2f}"],
        ["ITC at Risk (\u20b9)", f"\u20b9{s['itc_at_risk']:,.2f}"],
    ]

    for row_data in summary_data:
        ws_sum.append(row_data)

    ws_sum["A1"].font = Font(size=14, bold=True, color=HEADER)
    ws_sum["A4"].font = Font(bold=True, color="FFFFFF")
    ws_sum["A4"].fill = PatternFill("solid", fgColor=HEADER)
    ws_sum["B4"].font = Font(bold=True, color="FFFFFF")
    ws_sum["B4"].fill = PatternFill("solid", fgColor=HEADER)
    ws_sum["B15"].fill = PatternFill("solid", fgColor=RED)
    ws_sum["B15"].font = Font(bold=True)
    ws_sum.column_dimensions["A"].width = 42
    ws_sum.column_dimensions["B"].width = 28

    _write_sheet(wb, "\u2705 Matched", results["matched"], GREEN,
                 note="Invoices matched in both Purchase Register and GSTR-2A with no discrepancy.")
    _write_sheet(wb, "\u26a0 Amount Mismatch", results["amount_mismatch"], YELLOW,
                 note="Invoices found in both files but with different taxable/GST amounts. Review with supplier.")
    _write_sheet(wb, "\U0001f534 Missing in GSTR2A", results["missing_in_gstr2a"], RED,
                 note="In your Purchase Register but NOT filed by supplier. ITC cannot be claimed.")
    _write_sheet(wb, "\U0001f535 Not in Books", results["not_in_books"], BLUE,
                 note="In GSTR-2A (supplier filed) but missing from your Purchase Register.")

    return _save(wb), filename


def generate_ocr_report(df: pd.DataFrame) -> tuple[bytes, str]:
    """Invoice OCR extraction report. Returns (bytes, filename)."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Invoice_OCR_Report_{ts}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Invoices"

    if df.empty:
        ws.append(["No invoices found."])
        return _save(wb), filename

    ws.append(list(df.columns))
    _header_style(ws, row=1)

    for _, row in df.iterrows():
        ws.append([row[c] for c in df.columns])

    for i, row_vals in enumerate(df.itertuples(), start=2):
        fill_color = GREEN if getattr(row_vals, "confidence", "") == "high" else YELLOW
        fill = PatternFill("solid", fgColor=fill_color)
        for cell in ws[i]:
            cell.fill = fill

    _autofit(ws)
    return _save(wb), filename


def generate_payment_report(results: dict) -> tuple[bytes, str]:
    """Payment reconciliation report. Returns (bytes, filename)."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Payment_Recon_Report_{ts}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)
    s = results["summary"]

    ws_sum = wb.create_sheet("Summary", 0)
    ws_sum.sheet_view.showGridLines = False

    summary_rows = [
        ["GST PAYMENT RECONCILIATION", ""],
        [f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}", ""],
        ["", ""],
        ["Category", "Count / Value"],
        ["Total Liabilities",        s["total_liabilities"]],
        ["\u2705 Matched",           s["matched"]],
        ["\U0001f534 Unpaid",        s["unpaid"]],
        ["\u26a0\ufe0f Overpaid",   s["overpaid"]],
        ["\u26a0\ufe0f Underpaid",  s["underpaid"]],
        ["\u23f0 Late Payments",     s["late_payments"]],
        ["Unmatched Bank Entries",    s["unmatched_bank"]],
        ["", ""],
        ["Total Liability (\u20b9)", f"\u20b9{s['total_liability_amt']:,.2f}"],
        ["Total Paid (\u20b9)",      f"\u20b9{s['total_paid_amt']:,.2f}"],
        ["Outstanding (\u20b9)",     f"\u20b9{s['outstanding']:,.2f}"],
    ]
    for row_data in summary_rows:
        ws_sum.append(row_data)

    ws_sum["A1"].font = Font(size=14, bold=True, color=HEADER)
    ws_sum["A4"].font = Font(bold=True, color="FFFFFF")
    ws_sum["A4"].fill = PatternFill("solid", fgColor=HEADER)
    ws_sum["B4"].font = Font(bold=True, color="FFFFFF")
    ws_sum["B4"].fill = PatternFill("solid", fgColor=HEADER)
    if s["outstanding"] > 0:
        ws_sum["B15"].fill = PatternFill("solid", fgColor=RED)
        ws_sum["B15"].font = Font(bold=True)
    ws_sum.column_dimensions["A"].width = 36
    ws_sum.column_dimensions["B"].width = 28

    recon_df = results["reconciliation"]
    if not recon_df.empty:
        ws_r = wb.create_sheet("Reconciliation Detail")
        ws_r.append(list(recon_df.columns))
        _header_style(ws_r, row=1)
        for _, row in recon_df.iterrows():
            ws_r.append([str(v) if pd.isna(v) is False else "" for v in row])
        for i, row_vals in enumerate(recon_df.itertuples(), start=2):
            status = str(row_vals.status)
            fgc = GREEN if "Matched" in status else RED if "Unpaid" in status else YELLOW
            for cell in ws_r[i]:
                cell.fill = PatternFill("solid", fgColor=fgc)
        _autofit(ws_r)

    unmatched = results["unmatched_bank"]
    if not unmatched.empty:
        ws_u = wb.create_sheet("Unmatched Bank Entries")
        ws_u.append(["Entries in bank statement not matched to any GST liability", ""])
        ws_u.append([])
        ws_u.append(list(unmatched.columns))
        _header_style(ws_u, row=3)
        for _, row in unmatched.iterrows():
            ws_u.append(list(row))
        _fill_rows(ws_u, BLUE, start_row=4)
        _autofit(ws_u)

    return _save(wb), filename
